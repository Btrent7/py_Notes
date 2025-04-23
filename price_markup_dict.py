import pandas as pd
import openpyxl
from datetime import date

# Load Data Form
source_file = "C:/user/partNumber.xlsx"
df = pd.read_excel(source_file, sheet_name="pNum")

print(df)

# Extract values from specific cells
vendor_name = df.at[0, "Description"]       
vendor_sku = df.at[2, "Description"]         
item_name = df.at[3, "Description"]         
tpp_price = df.at[8, "Description"]     
markup_code = df.at[9, "Description"]    
site = df.at[10, "Description"].upper()
item_description = f"{vendor_name},#{vendor_sku},{item_name}".upper()
today = date.today().strftime("%m/%d/20%y")

# Markup rules
markup_rules = {
    '515C': lambda p: round((p * 1.5) / 1.2, 2),
    '300A': lambda p: round(p * 1.3, 2),
    '200C': lambda p: round((p * 1.2 + 5) / 1.1, 2),
    '600D': lambda p: round(p * 1.6),
}

def apply_markup(price, code):
    rule = markup_rules.get(code)
    return rule(price) if rule else "Error"

list_price = apply_markup(tpp_price, markup_code)

# Output to Data Table
output_file = "C:/user/partNumber2.xlsx"
df_output = pd.read_excel(output_file, sheet_name="table_pn")

print(df_output)

try:
    book = openpyxl.load_workbook(output_file)
    sheet = book["699_Table"]
except FileNotFoundError:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "699_Table"
    sheet.append(["New Part Number", "Vendor SKU", "Item Description", "Date", "TPP", "Marked-Up Price"])

# Find next blank row
next_row = sheet.max_row + 1

# Write data
sheet.cell(row=next_row, column=3).value = item_description
sheet.cell(row=next_row, column=2).value = vendor_sku
sheet.cell(row=next_row, column=5).value = tpp_price
sheet.cell(row=next_row, column=11).value = list_price
sheet.cell(row=next_row, column=4).value = today

# Save the file
book.save(output_file)
